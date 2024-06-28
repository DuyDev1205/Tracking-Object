from flask import Flask,render_template,jsonify, Response
from views import views
import cv2
import openpyxl
import time
from openpyxl.utils import get_column_letter, column_index_from_string
app = Flask(__name__)
data = []
count = 1
def generate_frames():
    camera = cv2.VideoCapture(0)  # Mở webcam

    while True:
        success, frame = camera.read()  # Đọc khung hình từ webcam
        if not success:
            break
        else:
            frame = cv2.flip(frame, 1)
            ret, buffer = cv2.imencode('.jpg', frame)
            frame = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
            

def read_excel_data(file_name,index):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    global data
    data = []
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=index, max_col=index+2):
        row_data = [cell.value for cell in row]
        data.append(row_data)
    return data

def getDate (file_name,index):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    new_column = get_column_letter(index)
    # Lấy giá trị của một ô cụ thể (ví dụ: ô B3)
    date = sheet[new_column+"1"].value
    print (new_column+"1")
    # Đóng tệp Excel
    workbook.close()
    return date

@app.route('/')
def index():
    global count
    excel_data = read_excel_data('test.xlsx',count)
    date = getDate('test.xlsx',count)
    return render_template('index.html', excel_data=excel_data, date=date)

@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/nextpage')
def nextpage():
    global count  # Khai báo rõ ràng là biến count là biến toàn cục
    if count < 52:
        count = count + 4
    excel_data = read_excel_data('test.xlsx',count)
    date = getDate('test.xlsx',count)
    return render_template('index.html', excel_data=excel_data, date=date)

@app.route('/prevpage')
def prevpage():
    global count  # Khai báo rõ ràng là biến count là biến toàn cục
    if count > 2:
        count = count - 4
    excel_data = read_excel_data('test.xlsx',count)
    date = getDate('test.xlsx',count)
    return render_template('index.html', excel_data=excel_data, date=date)
@app.route('/editAttend', methods=['POST'])
def geteditAttend():
    global data
    excel_data = read_excel_data('test.xlsx',count)

    return jsonify(excel_data)

def editAttend(name):
    global data
    modified_data = find_and_modify_name(name)
    write_excel_data( modified_data)

def write_excel_data(data):
    global count
    workbook = openpyxl.load_workbook("test.xlsx")
    sheet = workbook.active
    for row_index, row in enumerate(data, start=3):
        for col_index, value in enumerate(row, start=count):
            sheet.cell(row=row_index, column=col_index, value=value)
    workbook.save("test.xlsx")

def find_and_modify_name(name_to_find):
    global data
    for row in data:
        if row[1] == name_to_find:
            row[2] = 1
            break  # Khi tìm thấy và thay đổi giá trị, thoát khỏi vòng lặp
    return data

if __name__ == '__main__':
    app.run(debug = True)